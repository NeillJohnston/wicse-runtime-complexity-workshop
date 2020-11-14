#!/usr/bin/python3
import io
import re
import subprocess
import sys
import os
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import matplotlib.transforms as tnf

# Run a single-line command and return its stdout/stderr
def run(command, out=False, *args, **kwargs):
    output = subprocess.run(command, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, *args, **kwargs)
    stderr, stdout = output.stderr.decode('ascii'), output.stdout.decode('ascii')
    if out:
        print(stdout)
        print(stderr)
    return (stdout, stderr)

# Get the CPU time from a time -p string
def get_time(time_string):
    return float(time_string.splitlines()[1].split()[1]) + float(time_string.splitlines()[2].split()[1])

# Run a single response, store result in res table
def work(name, code):
    name_dir = f'subs/{name}'
    try:
        os.mkdir(name_dir)
    except:
        pass

    # Copy test files
    run(f'cp test/* {name_dir}')

    tests, _ = run(f'cd {name_dir} && ls *.in')
    tests = tests.splitlines()
    tests = [name.split('.')[0] for name in tests]
    tests = sorted(tests, key=int)

    # Write and compile code
    with open(f'{name_dir}/Main.java', 'w') as java:
        java.write(code)
    run(f'javac {name_dir}/Main.java')

    results = {}
    for test in tests:
        try:
            runs = 1
            for _ in range(runs):
                time = 0
                stdout, stderr = run(f'cd {name_dir} && timeout 10s time -p java Main < {test}.in | diff -Z {test}.out -')
                print(f'{name}, {test}, {len(stdout)}, {len(stderr)}')
                if len(stdout) == 0:
                    time += get_time(stderr)
                else:
                    return results

            length = int(test)
            results[length] = time/runs
        except:
            return results

    return results

# Set up empty chart
plt.xlabel('size of input (n)')
plt.ylabel('time to run (t)')
plt.plot([0], [0])
plt.savefig('plot.png', bbox_inches='tight')


# Load responses and run
res_file = sys.argv[1]
prob = sys.argv[2]
res_sheet = load_workbook(filename=res_file)[prob]
name_index, code_index = ('B', 'C') if res_sheet[f'B{1}'].value == 'Name' else ('C', 'B')
for i in range(2, res_sheet.max_row+1):
    name = res_sheet[f'{name_index}{i}'].value
    if name is None:
        break
    name = re.sub('[^a-zA-Z]', '_', name)
    code = res_sheet[f'{code_index}{i}'].value

    results = work(name, code)
    if len(results) == 0:
        print(f'{name} failed')
        continue

    n = sorted([n for n in results])
    t = [results[x] for x in n]
    plt.plot(n, t)
    plt.text(n[-1], t[-1], '      ' + name)
    plt.savefig('plot.png', bbox_inches='tight')
